"""Excel export service for generating reports."""
# type: ignore
# pyright: reportGeneralTypeIssues=false
from typing import List, Dict, Any
import openpyxl  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from io import BytesIO
from datetime import datetime


class ExcelExportService:
    """Service for generating Excel reports."""
    
    def __init__(self):
        """Initialize Excel export service."""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def generate_participant_report(
        self,
        participants: List[Dict[str, Any]],
        program_name: str
    ) -> BytesIO:
        """Generate participant list report."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Participants"
        
        # Add title
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = f"Participant Report - {program_name}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add metadata
        sheet['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        sheet['A3'] = f"Total Participants: {len(participants)}"
        
        # Headers
        headers = ['No', 'Name', 'Email', 'Phone', 'Status', 'Registration Date']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row_num, participant in enumerate(participants, 6):
            sheet.cell(row=row_num, column=1).value = row_num - 5
            sheet.cell(row=row_num, column=2).value = participant.get('name', '')
            sheet.cell(row=row_num, column=3).value = participant.get('email', '')
            sheet.cell(row=row_num, column=4).value = participant.get('phone', '')
            sheet.cell(row=row_num, column=5).value = participant.get('status', '')
            sheet.cell(row=row_num, column=6).value = participant.get('registration_date', '')
            
            # Apply border to all cells
            for col_num in range(1, 7):
                sheet.cell(row=row_num, column=col_num).border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, 7):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    
    async def generate_payment_report(
        self,
        payments: List[Dict[str, Any]],
        program_name: str,
        start_date: str,
        end_date: str
    ) -> BytesIO:
        """Generate payment report."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Payments"
        
        # Add title
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = f"Payment Report - {program_name}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add metadata
        sheet['A2'] = f"Period: {start_date} to {end_date}"
        sheet['A3'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        # Calculate totals
        total_amount = sum(p.get('amount', 0) for p in payments)
        sheet['A4'] = f"Total Amount: {total_amount:,.2f}"
        sheet['B4'] = f"Total Transactions: {len(payments)}"
        
        # Headers
        headers = ['No', 'Date', 'Participant', 'Amount', 'Payment Method', 'Status', 'Transaction ID', 'Reference']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=6, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row_num, payment in enumerate(payments, 7):
            sheet.cell(row=row_num, column=1).value = row_num - 6
            sheet.cell(row=row_num, column=2).value = payment.get('date', '')
            sheet.cell(row=row_num, column=3).value = payment.get('participant_name', '')
            sheet.cell(row=row_num, column=4).value = payment.get('amount', 0)
            sheet.cell(row=row_num, column=5).value = payment.get('payment_method', '')
            sheet.cell(row=row_num, column=6).value = payment.get('status', '')
            sheet.cell(row=row_num, column=7).value = payment.get('transaction_id', '')
            sheet.cell(row=row_num, column=8).value = payment.get('reference', '')
            
            # Format amount as currency
            sheet.cell(row=row_num, column=4).number_format = '#,##0.00'
            
            # Apply border
            for col_num in range(1, 9):
                sheet.cell(row=row_num, column=col_num).border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, 9):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 18
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    
    async def generate_custom_report(
        self,
        data: List[Dict[str, Any]],
        title: str,
        headers: List[str],
        sheet_name: str = "Report"
    ) -> BytesIO:
        """Generate a custom report with provided data."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        # Add title
        sheet.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        title_cell = sheet['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add metadata
        sheet['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        sheet['A3'] = f"Total Records: {len(data)}"
        
        # Headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row_num, record in enumerate(data, 6):
            for col_num, header in enumerate(headers, 1):
                # Use header as key (convert to lowercase and replace spaces with underscores)
                key = header.lower().replace(' ', '_')
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = record.get(key, '')
                cell.border = self.border
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
